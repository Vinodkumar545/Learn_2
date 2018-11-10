__author__ = "Vinodkumar Kouthal"

from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from selenium import webdriver
import unittest
import time

# Below are to record the LOG of every step, whenever the test is run.
import logging
import logging_conf
LOGGER = logging.getLogger("L2")

# Excel name, path and sheet name details
XL_NAME = 'test_data.xlsx'
XL_PATH = "/home/vinod/automation/Learn_2/" + XL_NAME
SHEET_NAME = "Test_Data"

# Excel Test Data Column IDs (for openpyxl, columns starts from 1)
COL_INP_1 = 1
COL_OPER = 2
COL_INP_2 = 3
COL_EXP = 4
COL_RES = 5

def super_calculator(driver, inp_1, oper, inp_2, exp_out, row_no):
	try:
		LOGGER.info("##########################################################################")

		# Enter the value/number in the first field
		driver.find_element_by_xpath("//input[@ng-model='first']").send_keys(inp_1)
		LOGGER.info(str("Value '" + str(inp_1) + "' is successfully inputed in input_1 field."))

		# Choose the operator from the drop down list (Select Method)
		select = Select(driver.find_element_by_xpath("//select[@ng-model='operator']"))
		select.select_by_visible_text(oper)
		LOGGER.info(str("Operation '" + str(oper) + "' is selected."))

		# Enter the value/number in the second field
		driver.find_element_by_xpath("//input[@ng-model='second']").send_keys(inp_2)
		LOGGER.info(str("Value '" + str(inp_2) + "' is successfully inputed in input_2 field."))

		# Click on the 'Go' button
		driver.find_element_by_xpath("//button[@id='gobutton']").click()
		LOGGER.info("Click on 'Go' button.")

		time.sleep(5) # Waiting for the output to be shown. 

		# capture the output
		super_calculator_output = driver.find_element_by_xpath("//h2[@class='ng-binding']").text
		LOGGER.info(str("Expected Value: " + str(exp_out) + " | Actual Value: " + str(super_calculator_output)))
		
		# Compare the expected and actual value to pass or fail a test case.
		if int(super_calculator_output) == exp_out:
			LOGGER.info("PASS: Value matches!")
			work_book[SHEET_NAME].cell(row=row_no, column=COL_RES).value = 'PASS'
			work_book.save(XL_NAME)
			return True
		else:
			LOGGER.info("FAIL: Value doesn't match!")
			work_book[SHEET_NAME].cell(row=row_no, column=COL_RES).value = 'FAIL'
			work_book.save(XL_NAME)
			return False
	except Exception as e: ## In case of any exception occur in try block execution
		LOGGER.debug(str(e))
		return False

def make_method(driver, inp_1, oper, inp_2, exp_out, row_no):

	def test_input(self):

		self.assertTrue(super_calculator(driver, inp_1, oper, inp_2, exp_out, row_no), u'{} {} {} != {} is NOT MATCH.'.format(inp_1, oper, inp_2, exp_out))

	# P.S. Ideal is, name should always be unique. If 2 data set has same inputs, then it would be consider as 1 test case
	test_input.__name__ = 'test_super_calculator_{}{}{}'.format(inp_1, oper, inp_2)

	return test_input

def add_methods(inputs):
	""" Take a TestCase and add a test method for each input """
	def decorator(klass):
		
		for inps in inputs:
			test_input = make_method(inps[0], inps[1], inps[2], inps[3], inps[4], inps[5])
			setattr(klass, test_input.__name__, test_input)

		return klass

	return decorator

def get_inputs():

	driver = webdriver.Chrome()
	driver.get("https://juliemr.github.io/protractor-demo/")
	driver.maximize_window()
	driver.implicitly_wait(10)
	assert 'Super Calculator' == driver.title

	global work_book
	work_book = load_workbook(XL_PATH)
	work_sheet = work_book[SHEET_NAME]
	tot_inputs = len(list(work_sheet.columns)[0])
	LOGGER.info(str("Total no. of test case: " + str(tot_inputs - 1)))

	input_query = []
	for row_no in range(2, tot_inputs + 1):
		individual_input = []
		
		individual_input.append(driver)
		individual_input.append(work_sheet.cell(row=row_no, column=COL_INP_1).value)
		individual_input.append(work_sheet.cell(row=row_no, column=COL_OPER).value)
		individual_input.append(work_sheet.cell(row=row_no, column=COL_INP_2).value)
		individual_input.append(work_sheet.cell(row=row_no, column=COL_EXP).value)
		individual_input.append(row_no) # Adding row_no, to write PASS/FAIL, once test case is validated.
		
		input_query.append(individual_input)

	LOGGER.info(str(input_query))
	return input_query

@add_methods(get_inputs())
class TestSuperCalculator(unittest.TestCase):
	pass
